import openpyxl
from django.contrib import admin
from django import forms
from django.http import HttpResponseRedirect
from .models import Group, Student, Lesson, Question, Result
from django.shortcuts import render
from django.urls import path
from django.core.cache import cache

class QuestionAdminForm(forms.ModelForm):

    class Meta:
        model = Question
        fields = "__all__"

        widgets = {
            "question": forms.Textarea(attrs={"rows":2}),
            "option_a": forms.TextInput(attrs={"size":80}),
            "option_b": forms.TextInput(attrs={"size":80}),
            "option_c": forms.TextInput(attrs={"size":80}),
            "option_d": forms.TextInput(attrs={"size":80}),
        }

class LessonAdmin(admin.ModelAdmin):
    list_display = ("title", "group")
    search_fields = ("title",)


class QuestionAdmin(admin.ModelAdmin):

    form = QuestionAdminForm

    list_display = ("question", "group", "lesson", "difficulty")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):

        if db_field.name == "group":

            if "last_group" in request.session:
                kwargs["initial"] = request.session["last_group"]

        if db_field.name == "lesson":

            if "last_lesson" in request.session:
                kwargs["initial"] = request.session["last_lesson"]

        return super().formfield_for_foreignkey(db_field, request, **kwargs)


    def response_add(self, request, obj, post_url_continue=None):

        request.session["last_group"] = obj.group_id
        request.session["last_lesson"] = obj.lesson_id
        request.session["last_difficulty"] = obj.difficulty

        return super().response_add(request, obj, post_url_continue)
        
    change_list_template = "admin/questions_changelist.html"

    def get_urls(self):
        urls = super().get_urls()

        custom_urls = [
            path(
                "import-excel/",
                self.import_excel
            ),
        ]

        return custom_urls + urls
    
    def import_excel(self, request):

        if request.method == "POST":
        
            # лимит импортов
            key = f"excel_import_{request.user.id}"
            count = cache.get(key, 0)

            if count >= 5:
                self.message_user(
                    request,
                    "Слишком много импортов. Подождите минуту.",
                    level="error"
                )
                return HttpResponseRedirect("../")
            cache.set(key, count + 1, timeout=60)
            
            print("FILES:", request.FILES)
            file = request.FILES.get("excel_file")

            if not file:
                self.message_user(request, "Файл не выбран!", level="error")
                return HttpResponseRedirect("../")

            # защита от больших файлов
            if file.size > 300 * 1024:
                self.message_user(request, "Файл слишком большой (макс 300Кб)", level="error")
                return HttpResponseRedirect("../")

            # разрешаем только xlsx
            if not file.name.endswith(".xlsx"):
                self.message_user(request, "Разрешены только Excel файлы (.xlsx)", level="error")
                return HttpResponseRedirect("../")

            try:
                wb = openpyxl.load_workbook(file)
            except Exception:
                self.message_user(request, "Ошибка чтения Excel файла", level="error")
                return HttpResponseRedirect("../")

            sheet = wb.active
            max_rows = sheet.max_row
            
            if max_rows != 16:
                self.message_user(
                    request,
                    f"Неверное кол-во строк ({max_rows}). Должно быть 16 (1 заголовок + 15 вопросов).",
                    level="error"
                )
                return HttpResponseRedirect("../")
                
            created = 0
            new_groups = 0
            new_lessons = 0
            
            diff_map = {
                "Легкий": 1,
                "Средний": 2,
                "Сложный": 3
            }

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

                try:

                    group_name, lesson_title, difficulty_text, question, a, b, c, d, correct = row

                    if not question:
                        continue

                    # --- проверка пустых полей ---
                    if not all([group_name, lesson_title, difficulty_text, question, a, b, c, d, correct]):
                        raise Exception(f"Пустые поля в строке {i}")
                    # --- защита длины вопроса ---
                    if len(str(question)) > 300:
                        raise Exception(f"Слишком длинный вопрос (строка {i})")
                    # --- защита длины ответов ---
                    for opt in [a,b,c,d]:
                        if len(str(opt)) > 200:
                            raise Exception(f"Слишком длинный вариант ответа (строка {i})")

                    if sheet.max_column != 9:
                        self.message_user(request, "Excel должен содержать ровно 9 колонок", level="error")
                        return HttpResponseRedirect("../")

                    # --- GROUP ---
                    group_name = str(group_name).strip()
                    group = Group.objects.filter(name=group_name).first()

                    if not group:
                        if new_groups >= 5:
                            raise Exception("Превышен лимит новых групп (макс 5 за импорт)")
                        group = Group.objects.create(name=group_name)
                        new_groups += 1

                    # --- LESSON ---
                    lesson_title = str(lesson_title).strip()
                    lesson = Lesson.objects.filter(
                        title=lesson_title,
                        group=group
                    ).first()

                    if not lesson:
                        if new_lessons >= 5:
                            raise Exception("Превышен лимит новых уроков (макс 5 за импорт)")
    
                        lesson = Lesson.objects.create(
                            title=lesson_title,
                            group=group
                        )
                        new_lessons += 1

                    difficulty = diff_map.get(str(difficulty_text).strip())

                    if not difficulty:
                        raise Exception(f"Неверная сложность в строке {i}")

                    # защита correct
                    if correct not in [1,2,3,4]:
                        raise Exception(f"Correct должен быть 1-4 (строка {i})")

                    Question.objects.create(
                        group=group,
                        lesson=lesson,
                        difficulty=difficulty,
                        question=str(question).strip(),
                        option_a=str(a),
                        option_b=str(b),
                        option_c=str(c),
                        option_d=str(d),
                        correct=int(correct)
                    )

                    created += 1

                except Exception as e:

                    self.message_user(
                        request,
                        f"Ошибка в строке {i}: {e}",
                        level="error"
                    )

            self.message_user(
                request,
                f"Импорт завершён. Добавлено вопросов: {created}"
            )

            return HttpResponseRedirect("../")

        return render(request, "admin/import_excel.html")

class GroupAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "order")
    list_editable = ("order",)
    ordering = ("order",)
  
admin.site.register(Group, GroupAdmin)
admin.site.register(Student)
admin.site.register(Lesson, LessonAdmin)
admin.site.register(Question, QuestionAdmin)
admin.site.register(Result)